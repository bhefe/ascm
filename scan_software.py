import winreg
import csv
import os
import sys
import socket
import getpass
import subprocess
import json
import re
import io
import urllib.parse
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink


# When running as frozen exe, PDFs are bundled in _MEIPASS temp dir
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
NOT_APPROVED_PDFS = [
    os.path.join(BASE_DIR, "LIST OF  SOFTWARES (NOT APPROVED).pdf")
]

# ── Pre-installed / system software to ignore (won't appear in report) ──
IGNORE_LIST = [
    "241916F58D6E7",
    "7.1 Surround Sound",
    "Actions Server",
    "Adobe Acrobat DC",
    "Adobe After Effects",
    "Adobe Genuine Service",
    "Adobe Media Encoder",
    "Adobe Photoshop",
    "Adobe Premiere Pro",
    "AI Noise Cancelation",
    "aimgr",
    "Antigravity (User)",
    "Application Compatibility Enhancements",
    "Armoury Crate",
    "ASUS Aac",
    "ASUS Ambient",
    "ASUS AURA",
    "ASUS Device Check",
    "ASUS Framework",
    "ASUS Hotplug",
    "ASUS Keyboard",
    "ASUS MB",
    "ASUS Mouse",
    "ASUS Smart",
    "ASUS Update",
    "ASUSAmbient",
    "ASUSPCAssistant",
    "AURA",
    "AVCEncoder Video Extension",
    "Bing Search",
    "Canon IJ Scan",
    "Cisco AnyConnect",
    "Clipchamp",
    "Composer - PHP",
    "Copilot",
    "Configuration Manager",
    "Cross Device",
    "Dev-C++",
    "DiagnosticsHub",
    "Dolby Vision",
    "DTSXUltra",
    "EA app",
    "EA SPORTS",
    "ELANTrack",
    "Entity Framework",
    "Epic Games",
    "EUC Wallpaper",
    "Game Assist",
    "Gaming App",
    "Gaming Services",
    "GameSDK",
    "Glanceby",
    "Glidex",
    "GlideX",
    "HPPrinter",
    "i Cloud",
    "InputMapper",
    "Intel Arc",
    "Intel Graphics",
    "Invgate Asset",
    "InvGate Insight",
    "It Takes Two",
    "Ivanti Secure",
    "Kinect for Windows",
    "Language Experience",
    "Launcher Prerequisites",
    "Lightshot",
    "Lenovo Companion",
    "Lenovo Migration",
    "Lenovo Quick",
    "Lenovo Settings",
    "Lenovo System Update",
    "Lenovo Vantage",
    "M365Companions",
    "Messaging",
    "Microsoft .NET",
    "Microsoft 365 Apps",
    "Microsoft ASP.NET",
    "Microsoft Azure",
    "Microsoft Edge",
    "Microsoft Family",
    "Microsoft GameInput",
    "Microsoft Intune",
    "Microsoft Power BI",
    "Microsoft ODBC",
    "Microsoft OneNote",
    "Microsoft SQL Server Compact",
    "Microsoft Office Hub",
    "Microsoft OneDrive",
    "Microsoft Policy",
    "Microsoft Server Speech",
    "Microsoft Teams",
    "Microsoft Visual Studio",
    "MSTeams",
    "Mozilla Maintenance",
    "NVIDIA",
    "Office 16 Click-to-Run",
    "Office Push Notification",
    "One Connect",
    "One Note",
    "One Note Virtual",
    "OTA Dependencies",
    "Outlook For Windows",
    "Overcooked",
    "Parsec",
    "PersonalGatewayComponents",
    "POV-Ray",
    "Power Automate",
    "Portal",
    "Print3D",
    "PuTTY",
    "Pulse Application",
    "Pulse Secure",
    "Quick Assist",
    "Realtek Audio",
    "RefreshRateService",
    "Rockstar Games",
    "ROG Live Service",
    "SAP Business",
    "Screen Pad Master",
    "SketchUp",
    "Steam",
    "Sec Health UI",
    "Shell Extension",
    "SmartPSS",
    "Start Experiences",
    "Synaptics Control",
    "Synaptics Utilities",
    "Teams Machine-Wide",
    "Tesseract-OCR",
    "Thunderbolt Control",
    "Tools for .Net",
    "TypeScript SDK",
    "Ubisoft Connect",
    "Universal CRT Tools",
    "VcXsrv",
    "Visual Studio Build Tools",
    "Visual Studio Community",
    "vs_clickoncesigntoolmsi",
    "vs_communitysharedmsi",
    "vs_filehandler_x86",
    "vs_minshellinteropmsi",
    "vs_minshellmsi",
    "vs_minshellmsires",
    "vs_minshellsharedmsi",
    "Web Experience",
    "Whiteboard",
    "WebStorm",
    "Widgets Platform",
    "Windows HDRCalibration",
    "Windows Subsystem For Linux",
    "WinRT Intellisense",
    "Wireshark",
    "WoJ XInput Emulator",
    "X-Rite Color",
    "Stable",
    "Update for",
    "Google Chrome",
    "Adobe Acrobat Reader Core App",
    "Free Snipping Tool",
    "Java Auto Updater",
    "Mozilla Firefox",
    "PL-2303 USB-to-Serial",
    "Printer Registration",
    "Synology Assistant",
    "TightVNC",
    "VMware vSphere Client",
    "Dell Support Assist",
    "Intel Optane",
    "Intel Wireless Bluetooth",
    "Maxx Audio",
    "Intel Trusted Connect",
    "Canon MF Scan",
    "LinkedIn",
    "Dell Digital Delivery",
    "Adobe Refresh Manager",
    "SAP",
    "AMD Radeon",
    "Lexmark",
    "EUC SECURITY",
    "INTEL INTERGRATED SENSOR",
    "Kaspersky",
    "TMBOX",
    "DefaultPackMSI",
    "EUC Security Update (2024-10)",
    "HP Notifications",
    "HPAudio Control",
    "HPSystem Information",
    "Intel Managementand Security Status",
    "Microsoft Search in Bing",
    "adobe acrobat",
    "microsoft project - en-us",
    "microsoft visio - en-us",
    "Java(TM) SE Development Kit 20",
]


def get_installed_software():
    """Scan Windows registry for installed software. Gracefully handles permission errors."""
    software_list = []
    success = True
    error_msg = None
    
    reg_paths = [
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall"),
        (winreg.HKEY_CURRENT_USER, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
    ]

    for hive, path in reg_paths:
        try:
            key = winreg.OpenKey(hive, path)
        except PermissionError:
            if success:  # Only set error on first occurrence
                success = False
                error_msg = "registry access denied (admin rights required)"
            continue
        except OSError as e:
            continue

        try:
            for i in range(winreg.QueryInfoKey(key)[0]):
                try:
                    subkey_name = winreg.EnumKey(key, i)
                    subkey = winreg.OpenKey(key, subkey_name)

                    try:
                        name = winreg.QueryValueEx(subkey, "DisplayName")[0]
                    except OSError:
                        name = None

                    try:
                        version = winreg.QueryValueEx(subkey, "DisplayVersion")[0]
                    except OSError:
                        version = ""

                    try:
                        publisher = winreg.QueryValueEx(subkey, "Publisher")[0]
                    except OSError:
                        publisher = ""

                    try:
                        install_date = winreg.QueryValueEx(subkey, "InstallDate")[0]
                    except OSError:
                        install_date = ""

                    if name and not _should_exclude(name, version, publisher):
                        software_list.append({
                            "Software": name.strip(),
                            "Version": str(version).strip(),
                            "Publisher": str(publisher).strip(),
                            "InstallDate": str(install_date).strip(),
                        })

                    subkey.Close()
                except OSError:
                    continue
        finally:
            key.Close()

    return software_list, success, error_msg


def _should_exclude(name: str, version: str, publisher: str) -> bool:
    """Filter out system components, drivers, and runtimes that SAM doesn't care about."""
    name_lower = name.lower()
    version_str = str(version).lower()
    publisher_lower = str(publisher).lower() if publisher else ""

    # Skip corrupted/invalid entries
    if name_lower in ("nan", "none", "") or name.startswith(("(", "{", "!")):
        return True

    # Skip entries that are just GUIDs
    if re.match(r'^[a-f0-9\-]{36}$|^[a-f0-9]{32}$', name_lower.strip()):
        return True

    # Skip Windows system components
    if any(x in name_lower for x in [
        "windows ", "win32 ", "system32", "drivers", "device", "helper",
        "sec health", "web experience", "widgets platform", "pc health check",
        "language experience pack", "keyboard layout", "ime",
        "update for microsoft", "microsoft update health",
    ]):
        return True

    # Skip drivers
    if any(x in name_lower for x in [
        " driver", " drivers", " mp drivers", " graphics", " chipset",
        "realtek", "nvidia", "intel graphics", "amd", "broadcom",
    ]):
        return True

    # Skip redistributables and runtimes
    if any(x in name_lower for x in [
        "visual c++", "vcredist", ".net runtime", ".net framework",
        "dotnet", "msvc", "redistributable", "runtime", "webrtc",
        "directx", "vulkan", "opengl", "dx12",
    ]):
        return True

    # Skip Office internal components
    if any(x in name_lower for x in [
        "office ", "click-to-run", " mui ", " setup metadata",
        "office shared", "office 16", "office push notification",
        "excel", "word", "powerpoint", "outlook", "onedrivesetup",
    ]):
        return True

    # Skip Microsoft internal utilities
    if any(x in name_lower for x in [
        "game assist", "cross device", "copilot", "portal",
        "power automate", "quick assist", "print3d",
        "glanceby", "intel graphics experience", "personal gateway",
    ]):
        return True

    # Skip evaluation/trial versions
    if "evaluation" in name_lower or "trial" in name_lower:
        return True

    return False


def get_store_apps():
    software_list = []

    SKIP_PREFIXES = {
        "microsoft.windows.", "microsoft.ui.", "microsoft.net.",
        "microsoft.vclibrary", "microsoft.services.", "microsoft.directx",
        "microsoft.designtool", "microsoft.advertising", "microsoft.xbox",
        "microsoft.getstarted", "microsoft.gethelp", "microsoft.people",
        "microsoft.wallet", "microsoft.storepurchase", "microsoft.desktopappinstaller",
        "microsoft.webp", "microsoft.heif", "microsoft.hevc", "microsoft.raw",
        "microsoft.webmedia", "microsoft.vp9video", "microsoft.av1video",
        "microsoft.mpeg2video", "microsoft.screenSketch", "microsoft.paint",
        "microsoft.windowscamera", "microsoft.windowsmaps", "microsoft.windowsalarms",
        "microsoft.windowscalculator", "microsoft.windowscommunicationsapps",
        "microsoft.windowsfeedbackhub", "microsoft.windowsnotepad",
        "microsoft.windowssoundrecorder", "microsoft.windowsstore",
        "microsoft.windowsterminal", "microsoft.zunemusic", "microsoft.zunevideo",
        "microsoft.bingweather", "microsoft.bingnews", "microsoft.bingfinance",
        "microsoft.microsoftsolitairecollection", "microsoft.microsoftstickynotes",
        "microsoft.photos", "microsoft.todos", "microsoft.family",
        "microsoft.cortana", "microsoft.onedrive", "microsoft.yourphone",
        "microsoft.phonelinkstub", "microsoft.549981c3f5f10", "microsoft.mspaint",
        "microsoft.powershell", "microsoft.screensketch", "inputapp",
        "1527c705-839a-4832-9118", "c5e2524a-ea46-4f67-841f",
        "e2a4f912-2574-4a75-9bb0", "f46d4000-fd22-4db4-ac8e",
        "microsoft.asynctextservice", "microsoft.ecapp", "apprex.",
        "realtekcontrolcenter", "nvidia.", "lenovoinc.", "lenovokbapp",
        "dolbylaboratories.",
    }

    KNOWN_APPS = {
        "spotifyab.spotifymusic": "Spotify", "5319275a.whatsapp": "WhatsApp",
        "facebook.instagram": "Instagram", "facebook.facebook": "Facebook",
        "4df9e0f8.netflix": "Netflix", "9426micro.twitter": "Twitter",
        "tiktok.tiktok": "TikTok", "telegramfz": "Telegram",
        "discord": "Discord", "slack": "Slack", "zoom.zoom": "Zoom",
        "clipchamp": "Clipchamp", "canva": "Canva", "amazon": "Amazon",
        "disney": "Disney+", "primevideo": "Prime Video", "linkedin": "LinkedIn",
        "todoist": "Todoist", "notion": "Notion", "viber": "Viber",
        "line": "LINE", "itunestoday": "iTunes",
    }

    ps_cmd = (
        'Get-AppxPackage | Where-Object {$_.IsFramework -eq $false -and $_.SignatureKind -ne \"System\"} '
        '| Select-Object Name, @{N=\"Version\";E={$_.Version}}, Publisher '
        '| ConvertTo-Json -Compress'
    )

    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps_cmd],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode != 0:
            return []

        data = json.loads(result.stdout)
        if isinstance(data, dict):
            data = [data]

        for app in data:
            raw_name = app.get("Name", "") or ""
            version = app.get("Version", "") or ""
            publisher_raw = app.get("Publisher", "") or ""

            if not raw_name:
                continue

            name_lower = raw_name.lower()

            if any(name_lower.startswith(prefix) for prefix in SKIP_PREFIXES):
                continue

            friendly = None
            for fragment, label in KNOWN_APPS.items():
                if fragment in name_lower:
                    friendly = label
                    break

            if not friendly:
                parts = raw_name.split(".")
                tail = parts[-1] if len(parts) > 1 else parts[0]
                friendly = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', tail)
                if not friendly.strip():
                    continue

            publisher = publisher_raw
            if publisher.startswith("CN="):
                publisher = publisher.split(",")[0].replace("CN=", "")

            software_list.append({
                "Software": friendly.strip(),
                "Version": str(version).strip(),
                "Publisher": publisher.strip(),
                "InstallDate": "",
            })

    except subprocess.TimeoutExpired:
        return software_list, False, "store apps scan timed out"
    except json.JSONDecodeError:
        return software_list, False, "store apps scan returned invalid data"
    except FileNotFoundError:
        return software_list, False, "powershell not found"
    except Exception as e:
        error_msg = str(e) if str(e) else "unknown error during store apps scan"
        return software_list, False, error_msg

    return software_list, True, None


def deduplicate_and_sort(software_list):
    seen = set()
    unique = []
    for entry in software_list:
        key = entry["Software"].lower()
        if key not in seen:
            seen.add(key)
            unique.append(entry)

    unique.sort(key=lambda x: x["Software"].lower())
    return unique


def extract_pdf_table(pdf_path):
    names = []
    sw_idx = None  # remember column index across pages
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                header = table[0]
                start = 1  # default: skip header row

                if sw_idx is None:
                    # First page: find the "Software" column
                    for i, col in enumerate(header):
                        col_name = (col or "").strip().lower()
                        if "software" in col_name:
                            sw_idx = i
                            break
                    if sw_idx is None:
                        sw_idx = 1 if len(header) > 1 else 0
                else:
                    # Continuation page: row 0 is data, not a header
                    start = 0

                for row in table[start:]:
                    if sw_idx < len(row):
                        val = (row[sw_idx] or "").strip().lower()
                        if val:
                            names.append(val)
    return names



APPROVED_LIST = [
    "putty",
    "postman",
    "vscode",
    "mysql workbench",
    "python",
    "mremoteng",
    "notepad ++",
    "pycharm",
    "node.js",
    "mysql server",
    "mysql shell",
    "android studio",
    "dbeaver",
    "heidi sql",
    "mongo db compass",
    "mongo db tools",
    "mongodb compass",
    "mongodb tools",
    "win scp",
    "vs code",
    "cursor",
    "tabby",
    "wps office",
    "go",
    "git",
    "postgres",
    "oracle vm virtual box",
    "ruby",
    "xampp",
    "nvm",
    "vim",
    "pdf sam enhanced",
    "mqtt explorer",
    "redis insight",
    "eclipse",
    "cygwin64",
    "ireport",
    "mariadb",
    "netbeans",
    "soapui",
    "tortoisegit",
    "winmerge",
    "microsoft visual studio code",
    "wsl kali linux",
    "kitty",
    "pycharm community edition",
    "pgadmin",
    "obs",
    "slack",
    "filezilla",
    "telegram",
    "minikube",
    "neo4j desktop",
    "zed attack proxy",
    "nmap",
    "ncap",
    "obsidian",
    "workbench",
    "sql developer",
    "zenmap",
    "libre",
    "sqlmap",
    "searchsploit",
    "feroxbuster",
    "bitwarden",
    "intellij idea community",
    "greenshot",
    "owasp",
    "teltonika com driver",
    "telematic configuration tool",
    "draw.io",
    "microsoft visual c++",
    "telegram desktop",
    "obs studio",
    "python 3",
    "pip",
    "flir ip config",
    "tesseract ocr",
    "oracle jinitiator",
    "snc client encryption",
    "7-zip",
    "adobe refresh manager",
    "apache directory studio",
    "application verifier",
    "arduino",
    "atom",
    "blender",
    "brave",
    "clickonce bootstrapper",
    "cmder",
    "discord",
    "eclipse mosquitto",
    "entity framework",
    "gns3",
    "gnuwin32",
    "go programming language",
    "google earth pro",
    "gpl ghostscript",
    "handbrake",
    "heidi",
    "helm",
    "hxd hex editor",
    "hydra",
    "icecap_collection",
    "inkscape",
    "intellij idea",
    "ivi shared components",
    "keepass",
    "keepassxc",
    "kicad",
    "kits configuration installer",
    "laragon",
    "mediatek sp driver",
    "microsoft .net framework",
    "microsoft .net host",
    "microsoft .net targeting pack",
    "microsoft gameinput",
    "microsoft netstandard sdk",
    "microsoft testplatform sdk",
    "microsoft windows desktop targeting pack",
    "miniforge",
    "mobile broadband hl service",
    "msi development tools",
    "mtk usb all",
    "mysql installer",
    "npgsql",
    "numbers",
    "ollama",
    "onlyoffice",
    "open vpn",
    "openlens",
    "opera",
    "oracle sql developer",
    "oracle virtualbox",
    "owasp zap",
    "pdf24 creator",
    "pdfsam basic",
    "photopea",
    "qgis",
    "quranflash",
    "r for windows",
    "remmina",
    "rstudio",
    "sd card formatter",
    "sdk arm64",
    "sdt commonssense tool",
    "soap ui",
    "sound lock",
    "sound recorder",
    "sqlcmd tools",
    "starship",
    "sticky note",
    "universal crt extension sdk",
    "universal crt headers",
    "universal crt redistributable",
    "universal general midi",
    "vagrant",
    "vcpp_crt",
    "visa.net shared components",
    "visual c++ library",
    "visual studio build tools 2022",
    "vlc media player",
    "vs immersive activate helper",
    "vs_blendmsi",
    "vs_clickoncebootstrappermsires",
    "vs_communitymsires",
    "vs_coreeditorfonts",
    "vs_devenvsharedmsi",
    "vs_filehandler_amd64",
    "vs_filetracker_singleton",
    "vs_githubprotocolhandlermsi",
    "vs_graphics_singletonx86",
    "vs_minshellinteropsharedmsi",
    "vs_sqlclickoncebootstrappermsi",
    "vs_tipsmsi",
    "vs_vswebprotocolselectormsi",
    "vscodium",
    "waterfox",
    "webtablet fb plugin",
    "wimgapi",
    "winappdeploy",
    "winbox64",
    "windows app certification kit",
    "windows iot extension sdk",
    "windows mobile extension sdk",
    "windows sdk",
    "windows_toolscorepkg",
    "winrt intellisense iot",
    "winrt intellisense ppi",
    "wireshark targeting pack",
    "wpt redistributables",
    "wptx64",
    "zoom",
    "sharex",
    "wacom tablet driver",
    "db browser for sqlite",
    "microsoft powerbi desktop",
    "power automate for desktop",
    "whatsapp",
    "winscp",
    "java 8",
    "openjdk",
    "wsl ubuntu",
    "github desktop",
    "ffmpeg",
    "audacity",
    "miktex",
    "texstudio",
    "laslook"
]

def build_official_list():
    entries = []
    seen = set()

    # Approved: use hardcoded list
    for name in APPROVED_LIST:
        clean = name.strip().lower()
        if clean and clean not in seen:
            entries.append((clean, "Allowed"))
            seen.add(clean)

    # Not Allowed: still read from PDF
    for f in NOT_APPROVED_PDFS:
        if os.path.isfile(f):
            for name in extract_pdf_table(f):
                if name not in seen:
                    entries.append((name, "Not Allowed"))
                    seen.add(name)

    return entries


def clean_name(name):
    if name is None:
        return ""
    name = str(name).lower()
    name = re.sub(r"\d+\.\d+[\.\d]*", "", name)
    name = re.sub(r"\(x64\)|\(64-bit\)|\(32-bit\)", "", name)
    name = re.sub(r"\(.*?\)", "", name)
    name = re.sub(r"[^\w\s]", " ", name)
    return " ".join(name.split()).strip()


def consolidate_software_name(name):
    """Extract base software name by removing versions, architecture, and components."""
    if name is None:
        return ""
    
    # Remove version numbers and architecture tags
    consolidated = re.sub(r"\s*\(32-bit\)|\s*\(64-bit\)|\s*\(x64\)|\s*\(x86\)", "", name, flags=re.IGNORECASE)
    consolidated = re.sub(r"\s*\d+\.\d+[\.\d]*", "", consolidated)
    
    # Remove common component suffixes
    suffixes = [
        r"\s*(core interpreter|development libraries|executables|standard library|pip bootstrap|runtime|library)",
        r"\s*(sdk|extension sdk|headers|headers and sources|extension)",
        r"\s*(development tools|build tools|setup metadata)",
        r"\s*\(user\)|\s*\(current user\)",
    ]
    
    for suffix in suffixes:
        consolidated = re.sub(suffix, "", consolidated, flags=re.IGNORECASE)
    
    # Clean up whitespace
    consolidated = " ".join(consolidated.split()).strip()
    return consolidated if consolidated else name


def run_check_local(csv_bytes):
    csv_rows = []
    for line in csv_bytes.decode("utf-8").splitlines(keepends=True):
        if not line.startswith("#"):
            csv_rows.append(line)

    user_reader = csv.DictReader(io.StringIO("".join(csv_rows)))

    official = build_official_list()
    official_names = [name for name, _ in official]
    official_status = [status for _, status in official]
    official_clean = [clean_name(name) for name in official_names]

    results = []
    counts = {"Allowed": 0, "Not Allowed": 0, "Not Found": 0}

    ignore_count = 0
    for row in user_reader:
        sw_name = str(row.get("Software") or "")

        # Skip if in ignore list (substring match) or too short to be real software
        if any(ignored.lower() in sw_name.lower() for ignored in IGNORE_LIST):
            ignore_count += 1
            continue
        if len(sw_name.strip()) <= 2:
            ignore_count += 1
            continue

        sw_clean = clean_name(sw_name)
        matched = ""
        status = "Not Found"

        for idx, off_clean in enumerate(official_clean):
            if not off_clean or not sw_clean:
                continue
            if len(sw_clean) < 3 or len(off_clean) < 3:
                continue
            if sw_clean in off_clean or off_clean in sw_clean:
                matched = official_names[idx]
                status = official_status[idx]
                break

        results.append({
            "software": sw_name,
            "status": status,
            "matched": matched,
        })
        counts[status] = counts.get(status, 0) + 1

    results.sort(key=lambda r: (["Not Allowed", "Not Found", "Allowed"].index(r["status"]), r["software"].lower()))
    
    # Consolidate similar software names (e.g., Python variants -> Python)
    consolidated = {}
    for r in results:
        base_name = consolidate_software_name(r["software"])
        key = (r["status"], base_name.lower())
        if key not in consolidated:
            consolidated[key] = {"software": base_name, "status": r["status"], "matched": r["matched"]}
    
    # Convert back to list and re-sort
    results = list(consolidated.values())
    results.sort(key=lambda r: (["Not Allowed", "Not Found", "Allowed"].index(r["status"]), r["software"].lower()))
    
    return results, counts


def main():
    # Force UTF-8 output to handle special characters in all terminals
    if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        except Exception:
            pass
    try:
        hostname = socket.gethostname()
        username = getpass.getuser()
        scan_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        print("\n" + "-"*40)
        print(f"  SOFTWARE COMPLIANCE SCAN V1.6")
        print(f"  Hostname: {hostname}")
        print(f"  Username: {username}")
        print(f"  Time: {scan_time}")
        print("-"*40 + "\n")

        # Step 1: SCAN
        print("  1. Scanning installed software...\n")
        registry_apps, reg_success, reg_error = get_installed_software()
        store_apps, store_success, store_error = get_store_apps()
        
        # Track data source status
        data_source_warnings = []
        if not reg_success:
            data_source_warnings.append(f"    ⚠ Registry scan failed: {reg_error}")
        if not store_success:
            data_source_warnings.append(f"    ⚠ Store apps scan failed: {store_error}")
        
        software = deduplicate_and_sort(registry_apps + store_apps)
        
        if data_source_warnings:
            print("  ⚠ Data Source Warnings:")
            for warning in data_source_warnings:
                print(warning)
            print()
        
        print(f"  Found {len(software)} installed programs.\n")

        # Step 2: Create CSV in memory for compliance check
        csv_output = io.StringIO()
        csv_output.write(f"# Hostname: {hostname}\n")
        csv_output.write(f"# Username: {username}\n")
        csv_output.write(f"# Scan Time: {scan_time}\n")
        writer = csv.DictWriter(csv_output, fieldnames=["Software", "Version", "Publisher", "InstallDate"])
        writer.writeheader()
        writer.writerows(software)
        csv_bytes = csv_output.getvalue().encode("utf-8")

        # Step 3: COMPARE with official list
        print("  2. Comparing against compliance lists...\n")
        results, counts = run_check_local(csv_bytes)

        # Step 4: SAVE EXCEL REPORT to Downloads
        safe_hostname = hostname.replace(" ", "_")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"compliance_report_{safe_hostname}_{timestamp}.xlsx"
        downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        os.makedirs(downloads_dir, exist_ok=True)
        output_path = os.path.join(downloads_dir, filename)

        # Generate Excel report
        generate_excel_report(output_path, results, counts, hostname, username, scan_time, data_source_warnings)
        
        # Step 5: DISPLAY RESULTS
        display_results(results, counts, hostname, data_source_warnings)
        


        print(f"  ")
        print(f"  Completed. Report saved to: {output_path}")
        
        return output_path

    except Exception as e:
        print(f"\n  [ERROR] {str(e)}", flush=True)
        import traceback
        traceback.print_exc()
        return None


def generate_excel_report(output_path, results, counts, hostname, username, scan_time, data_source_warnings=None):
    """Generate an Excel report with color coding and formatting."""
    if data_source_warnings is None:
        data_source_warnings = []
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Report"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    not_allowed_font = Font(color="FF0000", bold=True)
    not_found_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    allowed_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    row = 1
    
    # Header Section
    ws[f"A{row}"] = "Software Compliance Report"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    row += 1
    
    ws[f"A{row}"] = "Hostname:"
    ws[f"B{row}"] = hostname
    row += 1
    
    ws[f"A{row}"] = "Username:"
    ws[f"B{row}"] = username
    row += 1
    
    ws[f"A{row}"] = "Scan Time:"
    ws[f"B{row}"] = scan_time
    row += 1
    
    ws[f"A{row}"] = "Total Checked:"
    ws[f"B{row}"] = sum(counts.values())
    row += 2
    
    # Display data source warnings if any
    if data_source_warnings:
        warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        warning_font = Font(color="9C0006", bold=True)  # Dark red
        ws[f"A{row}"] = "⚠ Data Source Warnings:"
        ws[f"A{row}"].font = warning_font
        ws[f"A{row}"].fill = warning_fill
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        
        for warning in data_source_warnings:
            clean_warning = warning.replace("⚠ ", "").replace("    ", "")
            ws[f"A{row}"] = clean_warning
            ws[f"A{row}"].font = Font(color="9C0006")
            ws[f"A{row}"].fill = warning_fill
            ws.merge_cells(f"A{row}:C{row}")
            row += 1
        row += 1
    
    # NOT ALLOWED Section
    not_allowed = [r for r in results if r["status"] == "Not Allowed"]
    if not_allowed:
        ws[f"A{row}"] = f"NOT ALLOWED ({len(not_allowed)}) - Please uninstall immediately through IRIS helpdesk or contact IT unit"
        ws[f"A{row}"].font = not_allowed_font
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        
        ws[f"A{row}"] = "No."
        ws[f"B{row}"] = "Software name"
        for cell in [ws[f"A{row}"], ws[f"B{row}"]]:
            cell.font = Font(bold=True, color="000000")
            cell.border = border
            cell.alignment = center_align
        row += 1
        
        for idx, r in enumerate(not_allowed, 1):
            ws[f"A{row}"] = idx
            ws[f"B{row}"] = r["software"]
            for cell in [ws[f"A{row}"], ws[f"B{row}"]]:
                cell.font = not_allowed_font
                cell.border = border
                cell.alignment = left_align
            ws[f"A{row}"].alignment = center_align
            row += 1
        
        row += 1
    
    # ALLOWED Section
    allowed = [r for r in results if r["status"] == "Allowed"]
    if allowed:
        ws[f"A{row}"] = f"ALLOWED ({len(allowed)}) - Software's listed below requires SAM clearance memo. Pls email to: sam@tm.com.my"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        
        # Add clickable email link for ALLOWED section
        software_list = "\n".join([f"{idx}. {r['software']}" for idx, r in enumerate(allowed, 1)])
        body = f"Dear SAM,\n\nSoftware listed below requires clearance memo.\n\n{software_list}"
        subject = "Software Clearance Memo Request"
        mailto_url = f"mailto:sam@tm.com.my?subject={urllib.parse.quote(subject)}&body={urllib.parse.quote(body)}"
        
        ws[f"A{row}"] = "Click here to send clearance memo email"
        ws[f"A{row}"].font = Font(color="0563C1", underline="single")
        ws[f"A{row}"].hyperlink = Hyperlink(ref=f"A{row}", target=mailto_url)
        row += 1
        row += 1
        
        ws[f"A{row}"] = "No."
        ws[f"B{row}"] = "Software name"
        for cell in [ws[f"A{row}"], ws[f"B{row}"]]:
            cell.font = Font(bold=True, color="000000")
            cell.border = border
            cell.alignment = center_align
        row += 1
        
        for idx, r in enumerate(allowed, 1):
            ws[f"A{row}"] = idx
            ws[f"B{row}"] = r["software"]
            for cell in [ws[f"A{row}"], ws[f"B{row}"]]:
                cell.border = border
                cell.alignment = left_align
            ws[f"A{row}"].alignment = center_align
            row += 1
        
        row += 1
    
    # UNKNOWN Section
    not_found = [r for r in results if r["status"] == "Not Found"]
    if not_found:
        ws[f"A{row}"] = f"UNKNOWN ({len(not_found)}) - Software's listed below requires SAM clearance memo. Pls email to: sam@tm.com.my"
        ws[f"A{row}"].fill = not_found_fill
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        
        # Add clickable email link for UNKNOWN section
        software_list = "\n".join([f"{idx}. {r['software']}" for idx, r in enumerate(not_found, 1)])
        body = f"Dear SAM,\n\nSoftware listed below requires clearance memo.\n\n{software_list}"
        subject = "Software Clearance Memo Request"
        mailto_url = f"mailto:sam@tm.com.my?subject={urllib.parse.quote(subject)}&body={urllib.parse.quote(body)}"
        
        ws[f"A{row}"] = "Click here to send clearance memo email"
        ws[f"A{row}"].font = Font(color="0563C1", underline="single")
        ws[f"A{row}"].hyperlink = Hyperlink(ref=f"A{row}", target=mailto_url)
        row += 1
        row += 1
        
        ws[f"A{row}"] = "No."
        ws[f"B{row}"] = "Software name"
        for cell in [ws[f"A{row}"], ws[f"B{row}"]]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        row += 1
        
        for idx, r in enumerate(not_found, 1):
            ws[f"A{row}"] = idx
            ws[f"B{row}"] = r["software"]
            ws[f"A{row}"].border = border
            ws[f"A{row}"].alignment = center_align
            ws[f"B{row}"].border = border
            ws[f"B{row}"].alignment = left_align
            row += 1
    
    # Set column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    
    # Save the workbook
    wb.save(output_path)


def display_results(results, counts, hostname, data_source_warnings=None):
    """Display compliance results in a formatted table."""
    if data_source_warnings is None:
        data_source_warnings = []
    
    # Group results by status
    allowed = [r for r in results if r["status"] == "Allowed"]
    not_allowed = [r for r in results if r["status"] == "Not Allowed"]
    not_found = [r for r in results if r["status"] == "Not Found"]

    # Print summary
    print(f"  ")
    print(f"  +-- COMPLIANCE SUMMARY -----------------------+")
    print(f"  | Allowed:      {counts['Allowed']:>4} programs            |")
    print(f"  | Not Allowed:  {counts['Not Allowed']:>4} programs            |")
    print(f"  | Not Found:    {counts.get('Not Found', 0):>4} programs            |")
    print(f"  +---------------------------------------------+\n")
    
    # Print data source warnings if any
    if data_source_warnings:
        print(f"  ⚠ Data Source Warnings:")
        for warning in data_source_warnings:
            print(warning)
        print()

    # Print detailed results if there are issues
    if not_allowed:
        print(f"  ")
        print(f"\033[1;31m  -------------------------------------------------------------------------------------\033[0m")
        print(f"\033[1;31m  NOT ALLOWED ({len(not_allowed)}) Please uninstall immediately through IRIS helpdesk or contact IT unit\033[0m")
        print(f"\033[1;31m  -------------------------------------------------------------------------------------\033[0m")

        for r in not_allowed:
            print(f"\033[31m  | {r['software'][:60]}\033[0m")
        print(f"\033[31m  |\033[0m")

    if allowed:
        print(f"  ")
        print(f"  ----------------------------------------------------------------------------------------------")
        print(f"  ALLOWED ({len(allowed)}) Software's listed below requires SAM clearance memo. Pls email to: sam@tm.com.my")
        print(f"  ----------------------------------------------------------------------------------------------")

        for r in allowed:
            print(f"  | {r['software'][:60]}")

    if not_found:
        print(f"  = UNKNOWN ({len(not_found)}) =")
        for r in not_found[:20]:  # Show first 20
            print(f"  | {r['software'][:60]}")
        if len(not_found) > 20:
            print(f"  | ... and {len(not_found) - 20} more ...")

    if allowed and not not_allowed and not not_found:
        print(f"  [OK] All programs are compliant!\n")


if __name__ == "__main__":
    output_path = main()
    input("\n  Press Enter to close...")
    if output_path:
        try:
            os.startfile(output_path)
        except Exception as e:
            pass


